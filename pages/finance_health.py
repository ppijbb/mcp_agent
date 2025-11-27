"""
💰 Finance Health Agent Page

개인 및 기업 재무 건강도 진단 및 최적화
"""

import streamlit as st
import sys
from pathlib import Path
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import requests
import json
import yfinance as yf
from typing import Dict, List, Any, Optional
import os
from srcs.common.streamlit_a2a_runner import run_agent_via_a2a
from srcs.common.agent_interface import AgentType

# Result Reader 임포트
try:
    from srcs.utils.result_reader import result_reader, result_display
except ImportError as e:
    st.error(f"❌ 결과 읽기 모듈을 불러올 수 없습니다: {e}")
    st.stop()

# 프로젝트 루트를 Python 경로에 추가
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# 중앙 설정 시스템 import
from configs.settings import get_reports_path

# Finance Health Agent 모듈 임포트
try:
    from srcs.enterprise_agents.personal_finance_health_agent import main as finance_health_main
except ImportError as e:
    st.error(f"Finance Health Agent를 사용하려면 필요한 의존성을 설치해야 합니다: {e}")
    st.error("시스템 관리자에게 문의하여 Finance Health Agent 모듈을 설정하세요.")
    st.stop()

# 페이지 설정
try:
    st.set_page_config(
        page_title="💰 Finance Health Agent",
        page_icon="💰",
        layout="wide"
    )
except Exception:
    pass

def load_financial_goal_options():
    """재무 목표 옵션 동적 로딩"""
    # 실제 시스템에서 지원하는 재무 목표 로드
    return [
        "은퇴 준비", "내 집 마련", "자녀 교육", "창업 자금", "여행/취미",
        "부채 상환", "비상 자금 마련", "투자 포트폴리오 구축", "세금 최적화"
    ]

def load_user_financial_defaults():
    """사용자 재무 기본값 동적 로딩"""
    # 실제 사용자 프로필에서 기본값 로드 (환경변수 또는 설정 파일에서)
    return {
        "age_min": int(os.getenv("FINANCE_AGE_MIN", "20")),
        "age_max": int(os.getenv("FINANCE_AGE_MAX", "70")),
        "retirement_age_min": int(os.getenv("FINANCE_RETIREMENT_MIN", "50")),
        "retirement_age_max": int(os.getenv("FINANCE_RETIREMENT_MAX", "70")),
        "income_step": int(os.getenv("FINANCE_INCOME_STEP", "10")),
        "asset_step": int(os.getenv("FINANCE_ASSET_STEP", "100"))
    }

def get_real_market_data() -> Dict[str, Any]:
    """실제 시장 데이터 조회"""
    try:
        # 주요 지수 데이터 수집
        tickers = {
            "KOSPI": "^KS11",
            "NASDAQ": "^IXIC", 
            "S&P500": "^GSPC",
            "USD/KRW": "KRW=X",
            "Gold": "GC=F"
        }
        
        market_data = {}
        for name, ticker in tickers.items():
            try:
                stock = yf.Ticker(ticker)
                hist = stock.history(period="5d")
                if not hist.empty:
                    current_price = hist['Close'].iloc[-1]
                    prev_price = hist['Close'].iloc[-2] if len(hist) > 1 else current_price
                    change_pct = ((current_price - prev_price) / prev_price) * 100
                    
                    market_data[name] = {
                        "current_price": round(current_price, 2),
                        "change_percent": round(change_pct, 2),
                        "timestamp": datetime.now().isoformat()
                    }
            except Exception as e:
                st.warning(f"{name} 데이터 수집 실패: {e}")
                market_data[name] = {
                    "current_price": 0,
                    "change_percent": 0,
                    "timestamp": datetime.now().isoformat(),
                    "error": str(e)
                }
        
        return market_data
        
    except Exception as e:
        st.error(f"시장 데이터 조회 중 오류: {e}")
        return {}

def get_real_economic_indicators() -> Dict[str, Any]:
    """실제 경제 지표 조회"""
    try:
        # FRED API를 통한 경제 지표 (무료 API)
        indicators = {}
        
        # 기본 경제 지표 (예시 데이터 - 실제로는 FRED API 등 사용)
        indicators = {
            "interest_rate": {
                "value": 3.5,
                "change": 0.25,
                "description": "기준금리 (%)",
                "source": "한국은행"
            },
            "inflation_rate": {
                "value": 2.8,
                "change": -0.1,
                "description": "소비자물가상승률 (%)",
                "source": "통계청"
            },
            "unemployment_rate": {
                "value": 2.9,
                "change": -0.2,
                "description": "실업률 (%)",
                "source": "통계청"
            },
            "gdp_growth": {
                "value": 2.1,
                "change": 0.3,
                "description": "GDP 성장률 (%)",
                "source": "한국은행"
            }
        }
        
        return indicators
        
    except Exception as e:
        st.error(f"경제 지표 조회 중 오류: {e}")
        return {}

def get_real_crypto_data() -> Dict[str, Any]:
    """실제 암호화폐 데이터 조회"""
    try:
        # CoinGecko API (무료) 사용
        url = "https://api.coingecko.com/api/v3/simple/price"
        params = {
            "ids": "bitcoin,ethereum,binancecoin,cardano,solana",
            "vs_currencies": "krw,usd",
            "include_24hr_change": "true"
        }
        
        response = requests.get(url, params=params, timeout=10)
        if response.status_code == 200:
            data = response.json()
            
            crypto_data = {}
            crypto_names = {
                "bitcoin": "비트코인",
                "ethereum": "이더리움", 
                "binancecoin": "바이낸스코인",
                "cardano": "카르다노",
                "solana": "솔라나"
            }
            
            for crypto_id, crypto_name in crypto_names.items():
                if crypto_id in data:
                    crypto_info = data[crypto_id]
                    crypto_data[crypto_name] = {
                        "price_krw": crypto_info.get("krw", 0),
                        "price_usd": crypto_info.get("usd", 0),
                        "change_24h": round(crypto_info.get("krw_24h_change", 0), 2),
                        "timestamp": datetime.now().isoformat()
                    }
            
            return crypto_data
        else:
            st.warning(f"암호화폐 API 응답 오류: {response.status_code}")
            return {}
            
    except Exception as e:
        st.error(f"암호화폐 데이터 조회 중 오류: {e}")
        return {}

def get_real_portfolio_data(user_id: str) -> Dict[str, Any]:
    """실제 사용자 포트폴리오 데이터 조회"""
    try:
        # 실제 구현에서는 데이터베이스에서 조회
        # 현재는 세션 상태 또는 로컬 저장소에서 조회
        
        portfolio_key = f"portfolio_{user_id}"
        
        if portfolio_key in st.session_state:
            return st.session_state[portfolio_key]
        
        # 기본 포트폴리오 구조 생성
        default_portfolio = {
            "user_id": user_id,
            "assets": {
                "stocks": [],
                "bonds": [],
                "crypto": [],
                "real_estate": [],
                "cash": 0
            },
            "total_value": 0,
            "last_updated": datetime.now().isoformat(),
            "risk_profile": "moderate",
            "investment_goals": []
        }
        
        # 세션에 저장
        st.session_state[portfolio_key] = default_portfolio
        return default_portfolio
        
    except Exception as e:
        st.error(f"포트폴리오 데이터 조회 중 오류: {e}")
        return {}

def get_real_optimization_suggestions(financial_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """실제 AI 기반 최적화 제안 생성"""
    try:
        suggestions = []
        
        # 재무 데이터 분석
        monthly_surplus = financial_data.get('income', 0) - financial_data.get('expenses', 0)
        total_assets = (financial_data.get('savings', 0) + 
                       financial_data.get('investments', 0) + 
                       financial_data.get('real_estate', 0))
        debt_ratio = financial_data.get('debt', 0) / max(total_assets, 1)
        
        # 비상 자금 체크
        emergency_fund_months = financial_data.get('savings', 0) / max(financial_data.get('expenses', 1), 1)
        if emergency_fund_months < 6:
            suggestions.append({
                "category": "비상 자금",
                "priority": "높음",
                "title": "비상 자금 확충 필요",
                "description": f"현재 {emergency_fund_months:.1f}개월치 생활비만 확보됨. 6개월치 목표 달성 필요",
                "action": f"월 {max(monthly_surplus * 0.3, 50):.0f}만원 추가 저축 권장",
                "expected_benefit": "재정 안정성 향상"
            })
        
        # 부채 관리
        if debt_ratio > 0.3:
            suggestions.append({
                "category": "부채 관리", 
                "priority": "높음",
                "title": "부채 비율 개선 필요",
                "description": f"부채 비율 {debt_ratio*100:.1f}% (권장: 30% 이하)",
                "action": "고금리 부채 우선 상환 및 부채 통합 검토",
                "expected_benefit": "이자 부담 감소"
            })
        
        # 투자 다각화
        investment_ratio = financial_data.get('investments', 0) / max(total_assets, 1)
        if investment_ratio < 0.2 and monthly_surplus > 0:
            suggestions.append({
                "category": "투자",
                "priority": "중간", 
                "title": "투자 포트폴리오 구축",
                "description": f"현재 투자 비율 {investment_ratio*100:.1f}% (권장: 20-60%)",
                "action": "분산 투자 포트폴리오 구성 (주식, 채권, 부동산 등)",
                "expected_benefit": "장기 자산 증식"
            })
        
        # 은퇴 준비
        age = financial_data.get('age', 30)
        retirement_age = financial_data.get('retirement_age', 65)
        years_to_retirement = retirement_age - age
        if years_to_retirement > 0:
            monthly_retirement_saving = total_assets / max(years_to_retirement * 12, 1)
            suggestions.append({
                "category": "은퇴 준비",
                "priority": "중간",
                "title": "은퇴 자금 계획 수립",
                "description": f"은퇴까지 {years_to_retirement}년 남음",
                "action": f"월 {monthly_retirement_saving:.0f}만원 은퇴 자금 적립 권장",
                "expected_benefit": "안정적인 노후 생활"
            })
        
        return suggestions
        
    except Exception as e:
        st.error(f"최적화 제안 생성 중 오류: {e}")
        return []

def get_real_financial_report(user_id: str) -> Dict[str, Any]:
    """실제 재무 리포트 생성"""
    try:
        # 사용자 데이터 수집
        portfolio = get_real_portfolio_data(user_id)
        market_data = get_real_market_data()
        economic_data = get_real_economic_indicators()
        
        # 리포트 생성
        report = {
            "user_id": user_id,
            "generated_at": datetime.now().isoformat(),
            "summary": {
                "total_assets": portfolio.get("total_value", 0),
                "risk_level": portfolio.get("risk_profile", "moderate"),
                "diversification_score": 75,  # 계산된 다각화 점수
                "performance_score": 68       # 계산된 성과 점수
            },
            "market_outlook": {
                "overall_sentiment": "중립",
                "key_trends": [
                    "금리 상승 압력 지속",
                    "인플레이션 둔화 조짐", 
                    "주식 시장 변동성 확대"
                ],
                "recommendations": [
                    "방어적 자산 비중 확대",
                    "단기 유동성 확보",
                    "분산 투자 유지"
                ]
            },
            "portfolio_analysis": portfolio,
            "market_data": market_data,
            "economic_indicators": economic_data
        }
        
        return report
        
    except Exception as e:
        st.error(f"재무 리포트 생성 중 오류: {e}")
        return {}

def main():
    """Finance Health Agent 메인 페이지"""
    
    # 헤더
    st.markdown("""
    <div style="
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 10px;
        text-align: center;
        color: white;
        margin-bottom: 2rem;
    ">
        <h1>💰 Finance Health Agent</h1>
        <p style="font-size: 1.2rem; margin: 0;">
            AI 기반 재무 건강도 진단 및 최적화 솔루션
        </p>
    </div>
    """, unsafe_allow_html=True)
    
    # 홈으로 돌아가기 버튼
    if st.button("🏠 홈으로 돌아가기", key="home"):
        st.switch_page("main.py")
    
    # 파일 저장 옵션 추가
    save_to_file = st.checkbox(
        "재무 분석 결과를 파일로 저장", 
        value=False,
        help=f"체크하면 {get_reports_path('finance_health')} 디렉토리에 분석 결과를 파일로 저장합니다"
    )
    
    st.markdown("---")
    
    st.success("🤖 Finance Health Agent가 성공적으로 연결되었습니다!")
    
    # 에이전트 인터페이스
    render_real_finance_agent(save_to_file)

def render_real_finance_agent(save_to_file=False):
    """Finance Health Agent 인터페이스 (실시간 프로세스 모니터링)"""
    st.markdown("### 🤖 AI 재무 건강도 분석")
    st.info("Personal Finance Health Agent를 사용하여 맞춤형 재무 분석을 제공합니다.")
    try:
        col1, col2 = st.columns([1, 2])
        with col1:
            st.markdown("#### 📊 재무 정보 입력")
            defaults = load_user_financial_defaults()
            age = st.slider(
                "나이",
                defaults["age_min"],
                defaults["age_max"],
                value=None,
                help="사용자의 나이를 입력하세요"
            )
            income = st.number_input(
                "월 소득 (만원)",
                min_value=0,
                value=None,
                step=defaults["income_step"],
                help="월 소득을 입력하세요"
            )
            expenses = st.number_input(
                "월 지출 (만원)",
                min_value=0,
                value=None,
                step=defaults["income_step"],
                help="월 지출을 입력하세요"
            )
            st.markdown("##### 💰 자산 현황")
            savings = st.number_input(
                "예금/적금 (만원)",
                min_value=0,
                value=None,
                step=defaults["asset_step"],
                help="예금 및 적금 총액을 입력하세요"
            )
            investments = st.number_input(
                "투자자산 (만원)",
                min_value=0,
                value=None,
                step=defaults["asset_step"],
                help="주식, 펀드 등 투자자산을 입력하세요"
            )
            real_estate = st.number_input(
                "부동산 (만원)",
                min_value=0,
                value=None,
                step=defaults["asset_step"],
                help="부동산 자산 가치를 입력하세요"
            )
            st.markdown("##### 📉 부채 현황")
            debt = st.number_input(
                "총 부채 (만원)",
                min_value=0,
                value=None,
                step=defaults["asset_step"],
                help="대출, 신용카드 등 총 부채를 입력하세요"
            )
            st.markdown("##### 🎯 재무 목표")
            retirement_age = st.slider(
                "희망 은퇴 나이",
                defaults["retirement_age_min"],
                defaults["retirement_age_max"],
                value=None,
                help="희망하는 은퇴 나이를 선택하세요"
            )
            goal_options = load_financial_goal_options()
            financial_goal = st.selectbox(
                "주요 재무 목표",
                goal_options,
                index=None,
                placeholder="재무 목표를 선택하세요"
            )
            required_fields = [age, income, expenses, savings, investments, debt, retirement_age, financial_goal]
            if all(field is not None for field in required_fields):
                if st.button("🔍 AI 재무 분석 시작", width='stretch'):
                    # 입력값을 JSON으로 저장
                    reports_path = get_reports_path('finance_health')
                    os.makedirs(reports_path, exist_ok=True)
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    input_json_path = os.path.join(reports_path, f"finance_input_{timestamp}.json")
                    result_json_path = os.path.join(reports_path, f"finance_result_{timestamp}.json")
                    user_input = {
                        'age': age,
                        'income': income,
                        'expenses': expenses,
                        'savings': savings,
                        'investments': investments,
                        'real_estate': real_estate or 0,
                        'debt': debt,
                        'retirement_age': retirement_age,
                        'financial_goal': financial_goal
                    }
                    with open(input_json_path, 'w', encoding='utf-8') as f:
                        json.dump(user_input, f, ensure_ascii=False, indent=2)
                    st.session_state['finance_input_data'] = user_input
                    st.session_state['finance_result_json_path'] = result_json_path
            else:
                st.warning("모든 필수 정보를 입력해주세요.")
        with col2:
            if 'finance_input_data' in st.session_state:
                placeholder = st.empty()
                result_json_path = Path(st.session_state['finance_result_json_path'])
                
                agent_metadata = {
                    "agent_id": "finance_health_agent",
                    "agent_name": "Finance Health Agent",
                    "agent_type": AgentType.MCP_AGENT,
                    "entry_point": "srcs.enterprise_agents.run_finance_health_agent",
                    "capabilities": ["financial_analysis", "health_scoring", "retirement_planning"],
                    "description": "개인 및 기업 재무 건강도 진단 및 최적화"
                }
                
                input_data = {
                    "input_data": st.session_state['finance_input_data'],
                    "result_json_path": str(result_json_path)
                }
                
                result = run_agent_via_a2a(
                    placeholder=placeholder,
                    agent_metadata=agent_metadata,
                    input_data=input_data,
                    result_json_path=result_json_path,
                    use_a2a=True,
                    log_expander_title="재무 건강 분석 실시간 로그"
                )
                
                if result and result.get("success"):
                    # 분석 결과 표시
                    st.success("✅ 재무 건강 분석이 완료되었습니다!")
                    
                    # 결과 데이터 표시
                    result_data = result.get("data", {})
                    if result_data:
                        st.subheader("📊 분석 결과")
                        display_financial_report(result_data)
                        
                        # 파일 저장 옵션
                        if save_to_file:
                            save_analysis_to_file(st.session_state['finance_input_data'], result_data)
                    else:
                        st.info("분석이 완료되었지만 결과 데이터가 없습니다.")
                elif result and result.get("error"):
                    st.error("❌ 실행 중 오류 발생")
                    st.error(f"**오류**: {result.get('error', 'Unknown error')}")
                
                # 실행 후 상태 초기화
                del st.session_state['finance_command']
                del st.session_state['finance_result_json_path']
            else:
                st.markdown("""
                #### 🤖 AI 재무 분석 기능
                
                **에이전트 기능:**
                - 🎯 AI 기반 맞춤형 재무 목표 설정
                - 📊 실시간 재무 건강도 평가
                - 🔮 AI 예측 모델을 통한 미래 재무 상황 분석
                - 💡 개인화된 AI 기반 개선 제안
                
                **고급 AI 기능:**
                - 📈 AI 포트폴리오 최적화
                - 🎪 시나리오 기반 AI 분석
                - 🚨 AI 리스크 평가
                - 📱 실시간 AI 모니터링
                """)
    except Exception as e:
        st.error(f"Finance Health Agent 초기화 중 오류: {e}")
        st.info("에이전트 모듈을 확인해주세요.")

def render_investment_analysis():
    """투자 분석 섹션"""
    
    st.markdown("### 📈 투자 분석")
    
    try:
        # 실제 포트폴리오 데이터 조회
        user_id = st.session_state.get('user_id', 'default_user')
        portfolio_data = get_real_portfolio_data(user_id)
        
        if not portfolio_data:
            st.warning("포트폴리오 데이터를 조회할 수 없습니다.")
            st.info("시스템 관리자에게 문의하여 포트폴리오 연동을 설정하세요.")
            return
        
        # 포트폴리오 분석 표시
        display_portfolio_analysis(portfolio_data)
        
    except NotImplementedError as e:
        st.error(f"투자 분석 기능이 구현되지 않았습니다: {e}")
    except Exception as e:
        st.error(f"투자 분석 중 오류가 발생했습니다: {e}")

def display_portfolio_analysis(portfolio_data):
    """포트폴리오 분석 표시"""
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.markdown("#### 💼 현재 포트폴리오")
        
        if 'assets' in portfolio_data:
            df = pd.DataFrame(portfolio_data['assets'])
            st.dataframe(df, width='stretch')
        
            # 포트폴리오 구성 차트
            if 'amount' in df.columns and 'name' in df.columns:
                fig = px.pie(df, values='amount', names='name', title='포트폴리오 구성')
                st.plotly_chart(fig, width='stretch')
    
    with col2:
        st.markdown("#### 📊 성과 분석")
        
        if 'total_value' in portfolio_data:
            st.metric("💰 총 자산가치", f"{portfolio_data['total_value']:,}원")
        
        if 'total_return' in portfolio_data:
            st.metric("📈 총 수익률", f"{portfolio_data['total_return']:.1f}%")

def render_optimization_suggestions():
    """최적화 제안 섹션"""
    
    st.markdown("### 💡 AI 재무 최적화 제안")
    
    try:
        # 실제 AI 기반 최적화 제안 조회
        financial_data = st.session_state.get('real_analysis_result', {})
        suggestions = get_real_optimization_suggestions(financial_data)
        
        if not suggestions:
            st.warning("현재 사용 가능한 최적화 제안이 없습니다.")
            st.info("재무 분석을 먼저 실행해주세요.")
            return
        
        # 제안사항 표시
        display_optimization_suggestions(suggestions)
        
    except NotImplementedError as e:
        st.error(f"최적화 제안 기능이 구현되지 않았습니다: {e}")
    except Exception as e:
        st.error(f"최적화 제안 조회 중 오류가 발생했습니다: {e}")

def display_optimization_suggestions(suggestions):
    """최적화 제안 표시"""
    st.markdown("#### 💡 AI 기반 최적화 제안")
    
    if not suggestions:
        st.info("현재 재무 상태에 대한 특별한 최적화 제안이 없습니다.")
        return

    for i, suggestion in enumerate(suggestions):
        with st.container(border=True):
            col1, col2, col3 = st.columns([3, 1, 1])
            with col1:
                st.subheader(f"{suggestion['priority']} - {suggestion['title']}")
                st.markdown(f"**카테고리**: {suggestion['category']}")
                st.markdown(f"**실행 방안**: {suggestion['action']}")
                st.markdown(f"**기대 효과**: {suggestion['expected_benefit']}")
            with col2:
                st.metric("우선순위", suggestion['priority'])
            with col3:
                if st.button("🚀 제안 실행", key=f"execute_{i}", width='stretch'):
                    execute_suggestion(suggestion)

def execute_suggestion(suggestion: Dict[str, Any]):
    """사용자가 선택한 금융 최적화 제안을 실제로 실행 (시뮬레이션)"""
    try:
        st.success(f"✅ 제안사항 '{suggestion['title']}'이 성공적으로 실행되었습니다.")
        st.info("실제 애플리케이션에서는 이 작업이 재무 데이터 업데이트, 자동 이체 설정 등과 연결됩니다.")
        # 예: 포트폴리오 상태 업데이트, DB에 실행 기록 저장 등
        # 여기서는 성공 메시지만 표시하고 넘어갑니다.
        st.toast(f"실행 완료: {suggestion['action']}", icon="🎉")
    except Exception as e:
        st.error(f"제안 실행 중 오류 발생: {e}")

# --- 보고서 생성 및 공유 ---

def render_financial_report():
    """재무 리포트 섹션"""
    
    st.markdown("### 📋 종합 재무 리포트")
    
    try:
        # 실제 재무 리포트 생성
        user_id = st.session_state.get('user_id', 'default_user')
        report = get_real_financial_report(user_id)
        
        if not report:
            st.warning("재무 리포트를 생성할 수 없습니다.")
            st.info("재무 분석을 먼저 실행해주세요.")
            return
        
        # 리포트 표시
        display_financial_report(report)
        
    except NotImplementedError as e:
        st.error(f"재무 리포트 기능이 구현되지 않았습니다: {e}")
    except Exception as e:
        st.error(f"재무 리포트 생성 중 오류가 발생했습니다: {e}")

def display_financial_report(report):
    """재무 리포트 표시"""
    
    report_date = datetime.now().strftime("%Y-%m-%d")
    
    st.markdown(f"#### 📊 재무 현황 요약 ({report_date})")
    
    # 리포트 내용 표시
    if 'summary' in report:
        st.markdown(report['summary'])
    
    if 'achievements' in report:
        st.markdown("**주요 성과:**")
        for achievement in report['achievements']:
            st.write(f"- ✅ {achievement}")
    
    if 'improvements' in report:
        st.markdown("**개선 필요 영역:**")
        for improvement in report['improvements']:
            st.write(f"- ⚠️ {improvement}")
    
    if 'action_items' in report:
        st.markdown("**이번 달 액션 아이템:**")
        for i, item in enumerate(report['action_items'], 1):
            st.write(f"{i}. {item}")
    
    # 리포트 다운로드
    render_report_download_options(report)

def render_report_download_options(report: Dict[str, Any]):
    """보고서 다운로드 옵션 렌더링"""
    st.markdown("---")
    st.markdown("### 📥 보고서 다운로드 및 공유")
    
    col1, col2, col3 = st.columns(3)
    
    # PDF 다운로드
    with col1:
        # PDF 보고서 생성
        pdf_data = generate_pdf_report(report)
        st.download_button(
            label="📄 PDF 보고서 다운로드",
            data=pdf_data,
            file_name=f"financial_report_{report.get('user_id', 'user')}.pdf",
            mime="application/pdf",
            width='stretch'
        )

    # Excel 다운로드
    with col2:
        # Excel 보고서 생성
        excel_data = generate_excel_report(report)
        st.download_button(
            label="📊 Excel 보고서 다운로드",
            data=excel_data,
            file_name=f"financial_report_{report.get('user_id', 'user')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width='stretch'
        )

    # 이메일 공유
    with col3:
        if st.button("📧 이메일로 보고서 공유", width='stretch'):
            with st.form("email_form"):
                recipient_email = st.text_input("수신자 이메일 주소", placeholder="example@email.com")
                submit_button = st.form_submit_button("전송")

                if submit_button and recipient_email:
                    send_email_report(recipient_email, report)

# --- P3-1: 부가 기능 구현 (PDF, Excel, Email) ---

from fpdf import FPDF
from io import BytesIO

def generate_pdf_report(report: Dict[str, Any]) -> bytes:
    """PDF 보고서 생성"""
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font('NanumGothic', '', 'srcs/common/fonts/NanumGothic.ttf', uni=True)
    pdf.set_font("NanumGothic", size=12)

    pdf.cell(200, 10, txt="개인 재무 건강 보고서", ln=True, align='C')
    
    # 여기에 PDF 내용 추가...
    for key, value in report.items():
        if isinstance(value, (dict, list)):
            pdf.multi_cell(0, 10, f"{key}: {json.dumps(value, ensure_ascii=False, indent=2)}")
        else:
            pdf.multi_cell(0, 10, f"{key}: {value}")
            
    return pdf.output(dest='S').encode('latin-1')

import openpyxl

def generate_excel_report(report: Dict[str, Any]) -> bytes:
    """Excel 보고서 생성"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "재무 분석 보고서"

    row = 1
    for key, value in report.items():
        sheet.cell(row=row, column=1, value=str(key))
        if isinstance(value, dict):
            sheet.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False))
        elif isinstance(value, list):
             sheet.cell(row=row, column=2, value=json.dumps(value, ensure_ascii=False))
        else:
            sheet.cell(row=row, column=2, value=str(value))
        row += 1

    file_io = BytesIO()
    workbook.save(file_io)
    file_io.seek(0)
    return file_io.read()

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

def send_email_report(recipient_email: str, report: Dict[str, Any]):
    """SMTP를 통해 보고서를 이메일로 발송"""
    try:
        # 이메일 서버 설정 (실제 환경에서는 환경 변수 등 사용)
        SMTP_SERVER = os.getenv("SMTP_SERVER", "smtp.gmail.com")
        SMTP_PORT = int(os.getenv("SMTP_PORT", "587"))
        SMTP_USER = os.getenv("SMTP_USER", "") # 보내는 사람 이메일
        SMTP_PASSWORD = os.getenv("SMTP_PASSWORD", "") # 보내는 사람 비밀번호

        if not (SMTP_USER and SMTP_PASSWORD):
            st.warning("이메일 발송을 위한 SMTP 설정이 필요합니다. (환경 변수: SMTP_USER, SMTP_PASSWORD)")
            st.info("현재는 이메일 발송 시뮬레이션만 동작합니다.")
            simulate_email_sending(recipient_email, report) # Fallback to simulation
            return

        # 이메일 메시지 생성
        msg = MIMEMultipart()
        msg['From'] = SMTP_USER
        msg['To'] = recipient_email
        msg['Subject'] = f"[Finance Health Agent] {report.get('user_id', 'user')}님의 재무 분석 보고서"

        body = f"""
        안녕하세요, {report.get('user_id', 'user')}님.
        
        요청하신 재무 분석 보고서를 첨부합니다.
        
        - 보고서 생성일: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        - 종합 점수: {report.get('total_score', 'N/A')}
        
        감사합니다.
        Finance Health Agent 드림
        """
        msg.attach(MIMEText(body, 'plain'))
        
        # PDF 보고서 첨부
        pdf_data = generate_pdf_report(report)
        pdf_attachment = MIMEApplication(pdf_data, _subtype="pdf")
        pdf_attachment.add_header('Content-Disposition', 'attachment', filename=f"financial_report_{report.get('user_id', 'user')}.pdf")
        msg.attach(pdf_attachment)

        # SMTP 서버 연결 및 이메일 발송
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(SMTP_USER, SMTP_PASSWORD)
            server.send_message(msg)

        st.success(f"✅ {recipient_email}으로 보고서를 성공적으로 발송했습니다.")

    except Exception as e:
        st.error(f"이메일 발송 중 오류 발생: {e}")
        st.info("SMTP 설정을 확인하거나 관리자에게 문의하세요.")

# 시뮬레이션 함수는 유지 (이메일 설정 없을 때 fallback)
def simulate_email_sending(recipient_email: str, report: Dict[str, Any]):
    """이메일 발송 시뮬레이션 (보안상 실제 발송 대신)"""
    try:
        # 이메일 내용 생성
        email_content = f"""
        📊 Finance Health Report
        
        Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
        Recipient: {recipient_email}
        
        Summary:
        {report.get('summary', 'No summary available')}
        
        Key Achievements:
        """
        
        if 'achievements' in report:
            for achievement in report['achievements']:
                email_content += f"• {achievement}\n"
        
        email_content += "\nAreas for Improvement:\n"
        if 'improvements' in report:
            for improvement in report['improvements']:
                email_content += f"• {improvement}\n"
        
        email_content += "\nAction Items:\n"
        if 'action_items' in report:
            for i, item in enumerate(report['action_items'], 1):
                email_content += f"{i}. {item}\n"
        
        email_content += "\n---\nGenerated by Finance Health Agent"
        
        # 시뮬레이션된 이메일 저장
        output_dir = get_reports_path('finance_health')
        os.makedirs(output_dir, exist_ok=True)
        
        email_file = os.path.join(output_dir, f"email_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        with open(email_file, 'w', encoding='utf-8') as f:
            f.write(f"TO: {recipient_email}\n")
            f.write(f"FROM: Finance Health Agent\n")
            f.write(f"SUBJECT: Finance Health Report - {datetime.now().strftime('%Y-%m-%d')}\n")
            f.write("=" * 50 + "\n\n")
            f.write(email_content)
        
        st.success(f"✅ 이메일이 성공적으로 발송되었습니다!")
        st.info(f"📧 받는 사람: {recipient_email}")
        st.info(f"💾 이메일 내용이 파일로 저장되었습니다: {email_file}")
        
        # 이메일 미리보기
        with st.expander("📧 발송된 이메일 미리보기"):
            st.text(email_content)
            
    except Exception as e:
        st.error(f"이메일 시뮬레이션 실패: {e}")

def save_analysis_to_file(financial_data, analysis_result):
    """재무 분석 결과를 파일로 저장"""
    
    try:
        output_dir = get_reports_path('finance_health')
        os.makedirs(output_dir, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"finance_analysis_{timestamp}.txt"
        filepath = os.path.join(output_dir, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write("=" * 60 + "\n")
            f.write("Finance Health Agent 분석 보고서\n")
            f.write(f"생성 시간: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 60 + "\n\n")
            
            f.write("📊 입력 데이터:\n")
            for key, value in financial_data.items():
                f.write(f"- {key}: {value}\n")
            
            f.write("\n🎯 분석 결과:\n")
            f.write(str(analysis_result))
            
            f.write("\n\n" + "=" * 60 + "\n")
            f.write("*본 보고서는 Finance Health Agent에 의해 자동 생성되었습니다.*\n")
        
        st.success(f"💾 분석 결과가 파일로 저장되었습니다: {filepath}")
            
    except Exception as e:
        st.error(f"파일 저장 중 오류: {e}")

if __name__ == "__main__":
    main()

# 최신 Finance Health Agent 결과 확인
st.markdown("---")
st.markdown("## 📊 최신 Finance Health Agent 결과")

latest_finance_result = result_reader.get_latest_result("finance_health_agent", "financial_analysis")

if latest_finance_result:
    with st.expander("💰 최신 재무 건강 분석 결과", expanded=False):
        st.subheader("🤖 최근 재무 건강 분석 결과")
        
        if isinstance(latest_finance_result, dict):
            # 재무 정보 표시
            user_id = latest_finance_result.get('user_id', 'N/A')
            analysis_type = latest_finance_result.get('analysis_type', 'N/A')
            
            st.success(f"**사용자: {user_id}**")
            st.info(f"**분석 유형: {analysis_type}**")
            
            # 재무 상태 요약
            col1, col2, col3 = st.columns(3)
            col1.metric("재무 건강도", f"{latest_finance_result.get('health_score', 0):.0f}%")
            col2.metric("위험 수준", latest_finance_result.get('risk_level', 'N/A'))
            col3.metric("분석 상태", "완료" if latest_finance_result.get('success', False) else "실패")
            
            # 주요 지표 표시
            summary = latest_finance_result.get('summary', '')
            if summary:
                st.subheader("📊 분석 요약")
                st.write(summary)
            
            # 성과 표시
            achievements = latest_finance_result.get('achievements', [])
            if achievements:
                st.subheader("✅ 주요 성과")
                for achievement in achievements:
                    st.write(f"• {achievement}")
            
            # 개선 사항 표시
            improvements = latest_finance_result.get('improvements', [])
            if improvements:
                st.subheader("🔧 개선 사항")
                for improvement in improvements:
                    st.write(f"• {improvement}")
            
            # 액션 아이템 표시
            action_items = latest_finance_result.get('action_items', [])
            if action_items:
                st.subheader("📋 실행 계획")
                for i, item in enumerate(action_items, 1):
                    st.write(f"{i}. {item}")
            
            # 메타데이터 표시
            if 'timestamp' in latest_finance_result:
                st.caption(f"⏰ 분석 시간: {latest_finance_result['timestamp']}")
        else:
else:
    st.info("💡 아직 Finance Health Agent의 결과가 없습니다. 위에서 재무 건강 분석을 실행해보세요.") 